from django.contrib.auth import get_user_model
from django.core.exceptions import ValidationError
from django.utils import timezone
from rest_framework import status, viewsets
from rest_framework.decorators import action
from rest_framework.parsers import MultiPartParser
from rest_framework.permissions import IsAuthenticated
from django.http import HttpResponse
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from rest_framework.response import Response
from rest_framework.views import APIView

from accounts.permissions import IsSuperAdmin, IsSuperAdminOrDeptAdmin, IsSuperAdminOrReadOnly
from .importers import (
    import_departments, import_rooms, import_sections,
    import_timetable_entries,
)
from .models import (
    Block, Campus, Department, Floor, Room, Section,
    TemporaryAllocation, TimetableEntry, Timeslot,
)
from .serializers import (
    BlockSerializer, CampusSerializer, DepartmentSerializer, FloorSerializer,
    FreeRoomQuerySerializer, RoomSerializer, SectionSerializer,
    TemporaryAllocationSerializer, TimetableEntrySerializer,
)
from .serializers import TimeslotSerializer
from rest_framework.permissions import IsAuthenticated
import json
from django.http import JsonResponse
from pathlib import Path
from .services import find_free_rooms, get_room_status_for_slot, recommend_best_room
from rest_framework.permissions import IsAuthenticated

User = get_user_model()


class CampusViewSet(viewsets.ModelViewSet):
    queryset = Campus.objects.all()
    serializer_class = CampusSerializer
    permission_classes = [IsSuperAdminOrReadOnly]


class BlockViewSet(viewsets.ModelViewSet):
    queryset = Block.objects.select_related("campus").all()
    serializer_class = BlockSerializer
    permission_classes = [IsSuperAdminOrReadOnly]
    filterset_fields = ["campus"]


class FloorViewSet(viewsets.ModelViewSet):
    queryset = Floor.objects.select_related("block").all()
    serializer_class = FloorSerializer
    permission_classes = [IsSuperAdminOrReadOnly]
    filterset_fields = ["block"]


class DepartmentViewSet(viewsets.ModelViewSet):
    queryset = Department.objects.all()
    serializer_class = DepartmentSerializer
    permission_classes = [IsSuperAdminOrReadOnly]
    search_fields = ["name", "code"]


class RoomViewSet(viewsets.ModelViewSet):
    queryset = Room.objects.select_related("floor__block__campus", "department").all()
    serializer_class = RoomSerializer
    permission_classes = [IsSuperAdminOrDeptAdmin]
    filterset_fields = ["floor", "floor__block", "department", "room_type", "status"]
    search_fields = ["room_number"]

    @action(detail=True, methods=["get"], url_path="status")
    def slot_status(self, request, pk=None):
        """Check this room's availability for a given day/start_time/end_time."""
        room = self.get_object()
        serializer = FreeRoomQuerySerializer(data=request.query_params)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data
        result = get_room_status_for_slot(room, data["day"], data["start_time"], data["end_time"])
        result["room"] = RoomSerializer(room).data
        if result.get("occupying_section"):
            result["occupying_section"] = SectionSerializer(result["occupying_section"]).data
        return Response(result)


class SectionViewSet(viewsets.ModelViewSet):
    queryset = Section.objects.select_related(
        "department",
        "permanent_room__floor__block",
    ).all()
    serializer_class = SectionSerializer
    permission_classes = [IsSuperAdminOrDeptAdmin]
    filterset_fields = ["department", "year"]


class TimeslotViewSet(viewsets.ModelViewSet):
    queryset = Timeslot.objects.all()
    serializer_class = TimeslotSerializer
    permission_classes = [IsSuperAdminOrReadOnly]


class TimetableEntryViewSet(viewsets.ModelViewSet):
    queryset = TimetableEntry.objects.select_related(
        "section",
        "section__department",
        "section__permanent_room__floor__block",
        "room__floor__block",
        "timeslot",
    ).all()
    serializer_class = TimetableEntrySerializer
    permission_classes = [IsSuperAdminOrDeptAdmin]
    filterset_fields = ["section", "room", "day", "activity_type"]


class TemporaryAllocationViewSet(viewsets.ModelViewSet):
    queryset = TemporaryAllocation.objects.select_related("room", "section", "requested_by").all()
    serializer_class = TemporaryAllocationSerializer
    permission_classes = [IsAuthenticated]
    filterset_fields = ["room", "section", "day", "status"]

    def perform_create(self, serializer):
        serializer.save(requested_by=self.request.user)

    @action(detail=True, methods=["post"], permission_classes=[IsSuperAdminOrDeptAdmin])
    def approve(self, request, pk=None):
        allocation = self.get_object()
        allocation.status = TemporaryAllocation.Status.APPROVED
        allocation.approved_by = request.user
        allocation.save()
        return Response(TemporaryAllocationSerializer(allocation).data)

    @action(detail=True, methods=["post"], permission_classes=[IsSuperAdminOrDeptAdmin])
    def reject(self, request, pk=None):
        allocation = self.get_object()
        allocation.status = TemporaryAllocation.Status.REJECTED
        allocation.approved_by = request.user
        allocation.save()
        return Response(TemporaryAllocationSerializer(allocation).data)


class FreeRoomSearchView(APIView):
    """GET /api/campus/free-rooms/?day=WED&start_time=10:00&end_time=12:00&..."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        serializer = FreeRoomQuerySerializer(data=request.query_params)
        serializer.is_valid(raise_exception=True)
        d = serializer.validated_data
        rooms = find_free_rooms(
            day=d["day"], start_time=d["start_time"], end_time=d["end_time"],
            campus_id=d.get("campus_id"), block_id=d.get("block_id"), floor_id=d.get("floor_id"),
            room_type=d.get("room_type"), min_capacity=d.get("min_capacity"),
            department_id=d.get("department_id"),
        )
        return Response(RoomSerializer(rooms, many=True).data)


class RecommendRoomView(APIView):
    """POST /api/campus/recommend-room/ — best-fit room for a temporary allocation request."""
    permission_classes = [IsSuperAdminOrDeptAdmin]

    def post(self, request):
        serializer = FreeRoomQuerySerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        d = serializer.validated_data
        room = recommend_best_room(
            day=d["day"], start_time=d["start_time"], end_time=d["end_time"],
            required_capacity=d.get("min_capacity"),
            preferred_department_id=d.get("department_id"),
            preferred_block_id=d.get("block_id"),
            room_type=d.get("room_type"),
        )
        if not room:
            return Response({"detail": "No suitable room available."}, status=status.HTTP_404_NOT_FOUND)
        return Response(RoomSerializer(room).data)


class ImportDepartmentsView(APIView):
    permission_classes = [IsSuperAdminOrDeptAdmin]
    parser_classes = [MultiPartParser]

    def post(self, request):
        uploaded_file = request.FILES.get("file")
        if not uploaded_file:
            return Response({"detail": "Please upload a file under the 'file' field."}, status=status.HTTP_400_BAD_REQUEST)
        try:
            count = import_departments(uploaded_file)
            return Response({"imported": count})
        except ValidationError as exc:
            return Response({"detail": exc.message_dict if hasattr(exc, "message_dict") else exc.messages or str(exc)}, status=status.HTTP_400_BAD_REQUEST)


class ImportRoomsView(APIView):
    permission_classes = [IsSuperAdminOrDeptAdmin]
    parser_classes = [MultiPartParser]

    def post(self, request):
        uploaded_file = request.FILES.get("file")
        if not uploaded_file:
            return Response({"detail": "Please upload a file under the 'file' field."}, status=status.HTTP_400_BAD_REQUEST)
        try:
            count = import_rooms(uploaded_file)
            return Response({"imported": count})
        except ValidationError as exc:
            return Response({"detail": exc.message_dict if hasattr(exc, "message_dict") else exc.messages or str(exc)}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as exc:
            # Return readable error for frontend instead of HTTP 500
            return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)


class ImportSectionsView(APIView):
    permission_classes = [IsSuperAdminOrDeptAdmin]
    parser_classes = [MultiPartParser]

    def post(self, request):
        uploaded_file = request.FILES.get("file")
        if not uploaded_file:
            return Response({"detail": "Please upload a file under the 'file' field."}, status=status.HTTP_400_BAD_REQUEST)
        try:
            count = import_sections(uploaded_file)
            return Response({"imported": count})
        except ValidationError as exc:
            return Response({"detail": exc.message_dict if hasattr(exc, "message_dict") else exc.messages or str(exc)}, status=status.HTTP_400_BAD_REQUEST)


class ImportTimetableEntriesView(APIView):
    permission_classes = [IsSuperAdminOrDeptAdmin]
    parser_classes = [MultiPartParser]

    def post(self, request):
        uploaded_file = request.FILES.get("file")
        if not uploaded_file:
            return Response({"detail": "Please upload a file under the 'file' field."}, status=status.HTTP_400_BAD_REQUEST)
        try:
            preview_only = str(request.data.get("preview", "false")).lower() in {"1", "true", "yes"}
            # optional timeslot labels filter (JSON array)
            timeslot_labels_raw = request.data.get("timeslot_labels")
            timeslot_labels = None
            if timeslot_labels_raw:
                import json
                try:
                    timeslot_labels = json.loads(timeslot_labels_raw) if isinstance(timeslot_labels_raw, str) else list(timeslot_labels_raw)
                except Exception:
                    timeslot_labels = None

            if preview_only:
                payload = import_timetable_entries(uploaded_file, preview_mode=True, timeslot_labels=timeslot_labels)
                return Response(payload)

            result = import_timetable_entries(uploaded_file, timeslot_labels=timeslot_labels)
            return Response({
                "imported": result.get("imported", 0),
                "skipped": result.get("skipped", 0),
                "message": "Timetable entries imported successfully.",
                "errors": result.get("errors", []),
            })
        except ValidationError as exc:
            return Response({"detail": exc.message_dict if hasattr(exc, "message_dict") else exc.messages or str(exc)}, status=status.HTTP_400_BAD_REQUEST)


class RoomChoicesView(APIView):
    """Return available room_type and status choices from the Room model."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        types = [{"value": t[0], "label": t[1]} for t in Room.RoomType.choices]
        statuses = [{"value": s[0], "label": s[1]} for s in Room.Status.choices]
        return Response({"room_types": types, "statuses": statuses})


class RoomsTemplateView(APIView):
    """Generate an XLSX template with dropdowns for room_type and status."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        types = [t[0] for t in Room.RoomType.choices]
        statuses = [s[0] for s in Room.Status.choices]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Rooms Template"

        headers = ["floor", "room_number", "room_type", "capacity", "department", "status", "has_projector", "has_smart_board", "is_computer_lab", "has_ac", "has_wifi"]
        ws.append(headers)

        # add some sample rows
        for i in range(1, 21):
            ws.append([2, f"LHA{200 + i}", types[0], 60, "", statuses[0], True, False, False, True, True])

        # Create a hidden sheet with lists for validation
        list_ws = wb.create_sheet(title="Lists")
        # room types in column A
        for idx, val in enumerate(types, start=1):
            list_ws.cell(row=idx, column=1, value=val)
        # statuses in column B
        for idx, val in enumerate(statuses, start=1):
            list_ws.cell(row=idx, column=2, value=val)

        # Define data validation referencing the Lists sheet
        max_row = 1000
        dv_type = DataValidation(type="list", formula1="=Lists!$A$1:$A${}".format(len(types)), allow_blank=True)
        dv_status = DataValidation(type="list", formula1="=Lists!$B$1:$B${}".format(len(statuses)), allow_blank=True)

        # apply to the room_type column (D) and status column (G)
        ws.add_data_validation(dv_type)
        ws.add_data_validation(dv_status)
        for r in range(2, max_row + 1):
            dv_type.add(ws.cell(row=r, column=4))
            dv_status.add(ws.cell(row=r, column=7))

        # hide the Lists sheet
        list_ws.sheet_state = 'hidden'

        # save workbook to bytes
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        resp = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = 'attachment; filename=rooms-template.xlsx'
        return resp


class TimetableTemplateView(APIView):
    """Return a JSON template of days and timeslot columns used by the frontend grid."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        # Prefer an admin-editable DB template if present, otherwise fall back to the JSON file
        try:
            from .models import TimetableTemplate
            tpl = TimetableTemplate.objects.filter(active=True).order_by("-updated_at").first()
            if tpl:
                return Response({"days": tpl.days, "timeslots": tpl.timeslots})
        except Exception:
            # ignore DB errors and fall back to file
            pass

        base = Path(__file__).resolve().parent
        p = base / "timetable_template.json"
        try:
            with p.open("r", encoding="utf-8") as fh:
                data = json.load(fh)
        except Exception as exc:
            return Response({"detail": f"Could not load timetable template: {str(exc)}"}, status=500)
        return Response(data)


class DashboardStatsView(APIView):
    """GET /api/campus/dashboard/ — summary cards for the ERP-style dashboard."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        now = timezone.localtime()
        today_code = now.strftime("%a").upper()[:3]

        total_rooms = Room.objects.count()
        occupied = Room.objects.filter(status=Room.Status.OCCUPIED).count()
        maintenance = Room.objects.filter(status=Room.Status.MAINTENANCE).count()
        available = total_rooms - occupied - maintenance

        return Response({
            "total_blocks": Block.objects.count(),
            "total_floors": Floor.objects.count(),
            "total_rooms": total_rooms,
            "occupied_rooms": occupied,
            "available_rooms": max(available, 0),
            "todays_classes": TimetableEntry.objects.filter(
                day=today_code, activity_type=TimetableEntry.ActivityType.LECTURE
            ).count(),
            "todays_labs": TimetableEntry.objects.filter(
                day=today_code, activity_type=TimetableEntry.ActivityType.LAB
            ).count(),
            "temporary_allocations_today": TemporaryAllocation.objects.filter(
                day=today_code, status=TemporaryAllocation.Status.APPROVED
            ).count(),
            "pending_requests": TemporaryAllocation.objects.filter(
                status=TemporaryAllocation.Status.PENDING
            ).count(),
            "total_departments": Department.objects.count(),
            "total_sections": Section.objects.count(),
        })
